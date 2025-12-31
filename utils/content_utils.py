import re
from openpyxl import load_workbook
import json

def read_content(path):
    with open(path, "r", encoding="utf-8") as f:
        content = f.read()

    content = content.replace("```json", "")
    content = content.replace("```", "");
    return content;



def remove_dot_bracket(text: str) -> str:
    """
    去掉 'Hamilton et al. [21]' 中的 '. [21]'。
    """
    # 匹配模式：句点 + 空格 + 方括号内数字
    pattern = r'\.\s*\[\d+\]'
    new_txt = re.sub(pattern, '', text)
    new_txt = new_txt.strip()
    new_txt = new_txt.replace('.', '')
    return new_txt

def read_excel(file_path:str) -> str:
    # 1. 打开 Excel 文件
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active  # 默认读取第一个工作表

    # 逐行读取数据并打印
    final_result = "";
    for row in ws.iter_rows():
        values = [cell.value for cell in row]  # 提取每个 Cell 的值
        final_result = final_result + "\n" + str(values);

    return final_result;




def write_str_to_file(file_path: str, content: str, mode: str = 'w', encoding: str = 'utf-8'):
    """
    将字符串写入文件

    参数:
    - file_path: 文件路径
    - content: 要写入的字符串
    - mode: 写入模式, 'w' 覆盖写入, 'a' 追加写入
    - encoding: 文件编码
    """
    try:
        with open(file_path, mode, encoding=encoding) as f:
            f.write(content)
        print(f"已成功写入文件: {file_path}")
    except Exception as e:
        print(f"写入文件时出错: {e}")


if __name__ == "__main__":
    normal_study = remove_dot_bracket("Kraus et al.");
    print(f"normal_study is {normal_study}");